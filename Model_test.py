import shutil
import pandas as pd
import numpy as np
from keras.layers import Dense, Flatten, Conv1D, MaxPooling1D, Input
from keras import backend as K
from openpyxl import Workbook
from keras.models import Model
import os
from sklearn.metrics import r2_score


def test_NH3_model_cnn_single(path, model_stru1, model_para1, test_path, optimizer, loss, num):
    # Loading data
    df = pd.read_pickle(path)
    x_test = np.expand_dims(df.values[:, 0:-1].astype(float), axis=2)  # Adding a one-dimensional axis
    y_test = df.values[:, -1] / num
    # Divide training set, test set
    print("Loading of data complete!")
    x_test = x_test
    y_test = y_test
    loaded_model = load_first_cnn(model_stru1, model_para1)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = predict_data(loaded_model, x_test, optimizer, loss)
    mean_error(predicted, y_test, num)
    show_scores(predicted, y_test)
    save_model_excel(predicted, y_test, test_path, num)
    print("Test complete!")

def load_first_cnn(model_stru2, model_para1):
    loaded_model = test_build_CNN_model()
    loaded_model.load_weights(model_para1)
    print("Loading of pre-trained models complete!")
    return loaded_model


def test_build_CNN_model():
    input1 = Input(shape=(187, 1))
    conv_layer1_1 = Conv1D(16, 3, strides=1, activation='relu')(input1)
    conv_layer1_2 = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_1)
    conv_layer1_3 = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_2)
    max_layer1 = MaxPooling1D(3)(conv_layer1_3)
    conv_layer2_1 = Conv1D(32, 3, strides=1, activation='relu')(max_layer1)
    conv_layer2_2 = Conv1D(32, 3, strides=1, activation='relu')(conv_layer2_1)
    conv_layer2_3 = Conv1D(32, 3, strides=1, activation='relu')(conv_layer2_2)
    max_layer2 = MaxPooling1D(3)(conv_layer2_3)
    flatten = Flatten()(max_layer2)
    f1 = Dense(1, activation='linear', name='prediction_one')(flatten)
    model = Model(outputs=f1, inputs=input1)
    model.summary()
    return model



# Custom metric function, determination factor R_Squares
def coeff_determination(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))


# Calculating the mean absolute error
def mean_error(predicted, y_test, num):
    predicted = np.reshape(predicted, len(predicted))
    y_test_size = y_test
    predicted = np.array(predicted)*num
    y_test_size = np.array(y_test_size)*num
    # result = np.mean(abs(predicted * num - y_test_size * num))
    result = np.mean(abs((predicted  - y_test_size ) / y_test_size)) * 100
    print("MAPE:{:.2f}%".format(result))
    return result



# Predicted data
def predict_data(model, X_test, optimizer, loss):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = model.predict(X_test)
    return predicted


# Save prediction results
def save_model_excel(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'Predicted concentration')
    ws.cell(1, 2, 'Real concentration')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, Y_test[i] * num)
    wb.save(name)
    print("Save value to finish!")


# Calculation of the decision factor
def show_scores(predicted, Y_test):
    r2_scores = r2_score(predicted, Y_test)
    print("R2:{:.4}".format(r2_scores))
    return r2_scores



def del_files(path):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


if __name__ == '__main__':

    optimizer = "adam"
    loss = "mean_squared_error"
    Test_nh3_result_data_path = "Test_NH3_results_data"

    del_files(Test_nh3_result_data_path)
    os.mkdir(Test_nh3_result_data_path)
    Pkl_data_path = "Test_Processed_NH3_data/NH3_average_removebd.pkl"
    result_values_file_name = Test_nh3_result_data_path + "/nh3-results.xlsx"
    model_para = "Result_nh3_cnn/nh3-para.h5"
    model_stru = "Result_nh3_cnn/nh3-stru.json"
    test_NH3_model_cnn_single(Pkl_data_path, model_stru, model_para, result_values_file_name, optimizer, loss, 10000)