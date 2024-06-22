import os
import random
import numpy as np
import xlwt
import re
import pandas as pd
from openpyxl import Workbook
import shutil
from pandas import DataFrame
# from Sin_reconstruction_method import Sin_reconstruction
from Sin_reconstruction_method_test import Sin_reconstruction

def read_single_component_data(path, newpath, txt_num):
    '''
    * read_Data.py file is placed under the same root directory as the read data folder.
    * path：Enter the read data folder path.
    * Read the data folder layout as shown in the example.
    * After rerunning to read the data, if you rerun to read, you need to delete the newly generated **_ok folder in the Data folder before starting the operation.
    '''
    os.mkdir(newpath)
    path = path
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        root_ = []
        dirs_ = []
        a = 0
        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        file_name_after = newpath + '\\' + file_name_list + '_average'
        if not os.path.exists(file_name_after):
            os.mkdir(file_name_after)
        for i in root_[1:]:
            file_name_after_ = file_name_after + '\\' + dirs__[a]
            if not os.path.exists(file_name_after_):
                os.mkdir(file_name_after_)
            txt_ = []
            for file_name_ in os.listdir(i):
                txt_.append(i + '\\' + file_name_)
            txt_num = txt_num  # 取多少文件
            txt_ = txt_[-txt_num:]
            num_all = []
            for txt_name in txt_:
                contents = []
                with open(txt_name, 'r') as f:
                    for line in f.readlines():
                        line = line.split('\n')
                        line = line[0].split('\t')
                        line = list(map(float, line))
                        contents.append(line)
                for content in contents:
                    num_all.append("%.4f" % (float(content[1]) / txt_num))
                if len(num_all) > len(contents):
                    for ii in range(len(num_all)):
                        if ii < len(contents):
                            num_all[ii] = "%.4f" % (num_all[ii] + float(num_all[ii + len(contents)]))
                        else:
                            num_all.pop()
                num_all = list(map(float, num_all))
                f.close()
                txt_name_after = newpath + '\\' + file_name_list + '_average\\' + dirs__[a] + "\\" + dirs__[a] + ".txt"
                with open(txt_name_after, "w") as ff:
                    for li in num_all:
                        ff.write(str(li) + "\n")
                ff.close()
            a += 1
        print(file_name_list, "Data reading completed！")
    print("All  data read completed！")


def del_files(path1):
    if os.path.exists(path1):
        shutil.rmtree(path1, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


def remove_bd(newpath):
    path = newpath
    nn_ = []
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        n_p = path + '\\' + file_name_list + '_removebd'
        nn_.append(n_p)
        os.mkdir(n_p)
        root_ = []
        dirs_ = []

        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        root_.pop(0)
        root__ = root_[-1]
        dirs___ = dirs__[-1]
        root_.pop()
        dirs__.pop()
        bd_name = root__ + '\\' + dirs___ + '.txt'
        for i in range(len(root_)):
            data = []
            file_name = root_[i] + '\\' + dirs__[i] + '.txt'
            file_name_ = n_p + '\\' + dirs__[i] + '.txt'
            with open(bd_name) as bd_f:
                bd_file = bd_f.read().split('\n')
            bd_f.close()
            with open(file_name, 'r+') as f:
                file = f.read().split('\n')
            f.close()
            del (bd_file[-1])
            del (file[-1])
            bd_file = list(map(float, bd_file))
            file = list(map(float, file))
            for i in range(len(bd_file)):
                i_num = "%.4f" % ((file[i]) / (bd_file[i]))
                data.append(i_num)
            with open(file_name_, 'w') as f_:
                f_.truncate(0)
                for ii in data:
                    f_.write(ii + "\n")
            f_.close()
    print('Successfully removed the backing!')
    return nn_


def writeinexcel(path, nn):
    lu = []
    path = path
    le_ = 0
    le = 0
    wb1 = xlwt.Workbook(encoding='utf-8')
    w1 = wb1.add_sheet('one')
    ipath_ = path[0]
    file_name_lists_ = []
    file_name_lists = []
    for file_name_ in os.listdir(ipath_):
        file_name_ = re.sub('\D', '', file_name_)
        b = list(file_name_)
        # b.pop()
        ans = "".join(map(str, b))
        file_name_lists_.append(ans)
    file_name_lists_ = list(map(int, file_name_lists_))
    file_name_lists_.sort()
    le = len(file_name_lists_)
    for le_i in range(le):
        w1.write(0, le_i + le_, file_name_lists_[le_i])
    le_ = le_ + len(file_name_lists_)
    for file_name_ in os.listdir(ipath_):
        file_name_lists.append(file_name_)

    # print(file_name_lists)
    # print(nn)
    file_name_lists.sort(key=lambda x: int(x[:-nn]))
    for i_a in range(len(file_name_lists)):
        path_ = ipath_ + '\\' + file_name_lists[i_a]
        ii = 1
        for line in open(path_, encoding='utf-8'):
            if line == '\n':
                continue
            else:
                w1.write(ii, i_a, float(line))
                ii += 1

    wb1.save(path[0] + ".xls")
    p1 = path[0] + ".xls"
    lu.append(p1)

    if len(path) > 1:
        ipath__ = path[1]
        le__ = 0
        lei = 0
        wb2 = xlwt.Workbook(encoding='utf-8')
        w2 = wb2.add_sheet('one')
        file_name_lists_ = []
        file_name_lists = []
        for file_name_ in os.listdir(ipath__):
            file_name_ = re.sub('\D', '', file_name_)
            b = list(file_name_)
            # b.pop()
            ans = "".join(map(str, b))
            file_name_lists_.append(ans)
        file_name_lists_ = list(map(int, file_name_lists_))
        file_name_lists_.sort()
        lei = len(file_name_lists_)
        for le_i in range(lei):
            w2.write(0, le_i + le__, file_name_lists_[le_i])
        le__ = le__ + len(file_name_lists_)
        for file_name_ in os.listdir(ipath__):
            file_name_lists.append(file_name_)
        file_name_lists.sort(key=lambda x: int(x[:-nn]))
        for i_a in range(len(file_name_lists)):
            path_ = ipath__ + '\\' + file_name_lists[i_a]
            ii = 1
            for line in open(path_, encoding='utf-8'):
                if line == '\n':
                    continue
                else:
                    w2.write(ii, i_a, float(line))
                    ii += 1
                    # print(line)
        wb2.save(path[1] + ".xls")
        p2 = path[1] + ".xls"
        lu.append(p2)

    return lu


def koumanbian(x):
    aa = []
    for i in x:
        zz1 = np.polyfit([i for i in range(len(i))], i, 3)  # 拟合
        pp1 = np.poly1d(zz1)
        # aa.append(np.log(i / (pp1([i for i in range(len(i))])))*2) # 先组合在差分的时候乘2
        aa.append(np.log(i / (pp1([i for i in range(len(i))]))))
    return np.array(aa)


def dif_NH3(path):
    # Weak characteristic absorption region 1
    a = 344  # 199
    b = 530  # 213

    data = pd.read_excel(path)
    columns = data.columns
    wb_all = Workbook()
    wb_all.create_sheet(index=0, title="all")
    ws_all = wb_all.active

    data_line_1 = data.iloc[a - 1:b, :]

    data_line_1 = np.array(data_line_1)

    data_all = data_line_1
    # print(data_all.shape)
    data_all = pd.DataFrame(data_all, columns=columns)
    data_all = data_all.T
    data_all = np.array(data_all)

    data_deal_all_line = koumanbian(data_all)
    deal_data_line_all = pd.DataFrame(data_deal_all_line)
    # print("deal_data_line_all",deal_data_line_all)
    columns_ = deal_data_line_all.columns
    # all_data
    l = len(columns_)
    i1 = 0
    for i in range(l):
        i1 += 1
        lie = deal_data_line_all[columns_[i]]
        for j in range(len(lie)):
            ws_all.cell(1, i + 1, i1)
            ws_all.cell(j + 2, i + 1, lie[j])
    for col_i in range(len(columns)):
        col = columns[col_i]
        ws_all.cell(col_i + 2, l + 1, col)
    ws_all.cell(1, l + 1, l + 1)
    path = path.split('.')[0]
    name = path + '.xlsx'
    wb_all.save(name)
    print("NH3 Data differential completion!")
    return name


def Extended_data(path3, nh3_spectrum_path):
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active

    data1 = pd.read_excel(path3)
    columns1 = data1.columns
    le = 0
    for i in range(data1.shape[1]):
        lie1 = data1[columns1[i]]
        data_1 = lie1
        ws.cell(1, i + 1, str(columns1[i]))
        for ii in range(len(data_1)):
            ws.cell(ii + 2, i + 1, data_1[ii])
    le += int(data1.shape[0])
    index1 = data1.index
    for time in range(20):
        a = random.uniform(0, 0.5)
        a = round(a, 3)
        b = random.uniform(0, 0.5)
        b = round(b, 3)
        if a == 0 or b == 0:
            a += 0.1
            b += 0.1
        for i in range(data1.shape[0]):
            for j in range(data1.shape[0]):
                ind1 = list(data1.loc[index1[i]])
                # ind1.pop(0)
                ind2 = list(data1.loc[index1[j]])
                # ind2.pop(0)
                ind1 = np.array(ind1)
                ind2 = np.array(ind2)
                '''Combination Rules'''
                a = float(a)
                b = float(b)
                data_1 = a * ind1 + b * ind2
                for iii in range(len(data_1)):
                    ws.cell(le + 2 + j, iii + 1, data_1[iii])
            le += int(data1.shape[0])
        print("circulate", time, "time")
    wb.save(nh3_spectrum_path)
    path1 = to_pkl(nh3_spectrum_path)
    print("NH3 Data expansion completed!")
    return path1


def to_pkl(path):
    # Read excel files
    df1 = DataFrame(pd.read_excel(path))
    dir_name = os.path.dirname(path)
    base_name = os.path.basename(path)
    suffix = base_name.split(".")[0]
    path_ = dir_name + "/" + suffix + ".pkl"
    df1.to_pickle(path_)
    return path_


def UV_DOAS():
    Raw_data_path = "Raw_NH3_data"
    Processed_data_path = "Processed_NH3_data"
    nh3_spectrum_path = Processed_data_path + '/nh3-spectrum.xlsx'
    txt_num = 50
    numb = 10
    del_files(Processed_data_path)
    read_single_component_data(Raw_data_path, Processed_data_path, txt_num)
    path1 = remove_bd(Processed_data_path)
    path2 = writeinexcel(path1, numb)
    path3 = dif_NH3(path2[0])
    # path3 = "Processed_NH3_data\\NH3_average_removebd.xlsx"
    path5 = Extended_data(path3,
                          nh3_spectrum_path)  # Concentration needs to be changed to real concentration before expanding data.
    Sin_reconstruction(path5)

def Do_Test_data():
    Raw_data_path = "Test_data"
    Processed_data_path = "Test_Processed_NH3_data"
    nh3_spectrum_path = Processed_data_path + '/nh3-spectrum.xlsx'
    txt_num = 50
    numb = 10
    del_files(Processed_data_path)
    read_single_component_data(Raw_data_path, Processed_data_path, txt_num)
    path1 = remove_bd(Processed_data_path)
    path2 = writeinexcel(path1, numb)
    path3 = dif_NH3(path2[0])
    # path3 = "Processed_NH3_data\\NH3_average_removebd.xlsx"
    path5 = Extended_data(path3, nh3_spectrum_path)  # Concentration needs to be changed to real concentration before expanding data.
    # path3 = "Test_Processed_NH3_data\\NH3_average_removebd.xlsx"
    path3=to_pkl(path3)
    Sin_reconstruction(path3)

if __name__ == '__main__':
    UV_DOAS()
    # Do_Test_data()