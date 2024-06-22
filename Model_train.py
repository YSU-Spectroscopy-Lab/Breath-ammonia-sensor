import Model_build
import os
import shutil
from keras import backend as K
from openpyxl import Workbook
import pandas as pd
import numpy as np
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt


def train_model(model, name, loss_name, model_para, model_stru, optimizer, loss, X_train, X_test, Y_train, Y_test,
                name1, name2, MAPE_r2_name):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])

    history = model.fit(X_train, Y_train,
                        batch_size=256,
                        epochs=1000, #1000
                        validation_data=(X_test, Y_test)
                        )
    predicted = predict_data(model, X_test, optimizer=optimizer, loss=loss)
    MAPE = mean_error(predicted, Y_test, num)

    save_model(model, model_stru, model_para)
    r2 = show_scores(predicted, Y_test)
    save_model_excel(predicted, Y_test, name, num)
    plot_history(history, name1, name2)
    save_model_loss(history, loss_name)
    save_MAPE_r2(MAPE, r2, MAPE_r2_name)
    return MAPE



# Custom metric function, determination factor R_Squares
def coeff_determination(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))


# Save prediction results
def save_model_excel(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'Predicted concentration')
    ws.cell(1, 2, 'Real concentration')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, Y_test[i] * num)
    wb.save(name)
    print("Save value to finish!")


def save_model_loss(history, name):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    hist = pd.DataFrame(history.history)
    loss = hist['loss']
    val_loss = hist['val_loss']
    ws.cell(1, 1, 'loss')
    ws.cell(1, 2, 'val_loss')
    for i in range(len(loss)):
        ws.cell(i + 2, 1, loss[i])
        ws.cell(i + 2, 2, val_loss[i])
    wb.save(name)
    print("Save loss to finish!")


# Predicted data
def predict_data(model, X_test, optimizer, loss):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = model.predict(X_test)
    return predicted


# Calculation of the decision factor
def show_scores(predicted, Y_test):
    r2_scores = r2_score(predicted, Y_test)
    print("R2:", r2_scores)
    return r2_scores



# Calculating the mean absolute error
def mean_error(predicted, y_test, num):
    predicted = np.reshape(predicted, len(predicted))
    y_test_size = y_test
    predicted = np.array(predicted) * num
    y_test_size = np.array(y_test_size) * num
    # result = np.mean(abs(predicted * num - y_test_size * num))
    result = np.mean(abs((predicted - y_test_size) / y_test_size)) * 100
    print("MAPE:{:.2f}%".format(result))
    return result


# Preservation of models
def save_model(model, name1, name2):
    # Convert their model grid structure to json storage
    # Store model parameter weights as h5 files
    model_json = model.to_json()
    with open(name1, 'w') as json_file:
        json_file.write(model_json)
    model.save_weights(name2)
    print("Save model complete!")


def plot_history(history, name1, name2):
    hist = pd.DataFrame(history.history)
    hist['epoch'] = history.epoch
    plt.figure()
    plt.xlabel('Epoch')
    plt.ylabel('loss')
    plt.plot(hist['epoch'], hist['loss'],
             label='Train loss')
    plt.plot(hist['epoch'], hist['val_loss'],
             label='Val loss')
    plt.ylim([-0.001, 0.05])
    plt.legend()
    plt.savefig(name1, dpi=600)
    plt.figure()
    plt.xlabel('Epoch')
    plt.ylabel('coeff_determination')
    plt.plot(hist['epoch'], hist['coeff_determination'],
             label='Train coeff_determination')
    plt.plot(hist['epoch'], hist['val_coeff_determination'],
             label='Val coeff_determination')
    plt.ylim([-0.5, 1.5])
    plt.legend()
    plt.savefig(name2, dpi=600)
    # plt.show()


def save_MAPE_r2(MAPE, r2, name):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active

    ws.cell(1, 1, 'MAPE')
    ws.cell(1, 2, 'R2')
    ws.cell(2, 1, MAPE)
    ws.cell(2, 2, r2)
    wb.save(name)
    print("Save MAPE & r2 to finish!")


def del_files(path):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


if __name__ == '__main__':
    Result_nh3_data_path = "Result_nh3_cnn"

    optimizer = "adam"
    loss = "mean_squared_error"

    model_structure = Result_nh3_data_path + "/CNN.png"
    result_values_file_name = Result_nh3_data_path + "/nh3-results.xlsx"
    loss_file_name = Result_nh3_data_path + "/nh3-loss.xlsx"
    model_para = Result_nh3_data_path + "/nh3-para.h5"
    model_stru = Result_nh3_data_path + "/nh3-stru.json"
    model_loss = Result_nh3_data_path + "/nh3-loss.png"
    model_R2 = Result_nh3_data_path + "/nh3-R2.png"
    model_MAPE_r2_name = Result_nh3_data_path + "/nh3-MAPE-r2.xlsx"


    del_files(Result_nh3_data_path)
    os.mkdir(Result_nh3_data_path)
    model, X_train, X_test, Y_train, Y_test, num = Model_build.run(model_structure)
    MAPE = train_model(model, result_values_file_name, loss_file_name, model_para, model_stru, optimizer, loss, X_train,
                X_test, Y_train,
                Y_test, model_loss, model_R2, model_MAPE_r2_name)

