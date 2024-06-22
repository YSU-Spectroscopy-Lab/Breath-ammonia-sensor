import os
import random
import numpy as np
import xlwt
import re
import pandas as pd
from openpyxl import Workbook
import shutil
from pandas import DataFrame
import matplotlib.pyplot as plt
from pylab import mpl
from scipy import stats
import math
from scipy.optimize import curve_fit


def read_single_component_data(path, newpath, txt_num):
    '''
    * read_Data.py file is placed under the same root directory as the read data folder.
    * path：Enter the read data folder path.
    * Read the data folder layout as shown in the example.
    * After rerunning to read the data, if you rerun to read, you need to delete the newly generated **_ok folder in the Data folder before starting the operation.
    '''
    os.mkdir(newpath)
    path = path
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        root_ = []
        dirs_ = []
        a = 0
        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        file_name_after = newpath + '\\' + file_name_list + '_average'
        if not os.path.exists(file_name_after):
            os.mkdir(file_name_after)
        for i in root_[1:]:
            file_name_after_ = file_name_after + '\\' + dirs__[a]
            if not os.path.exists(file_name_after_):
                os.mkdir(file_name_after_)
            txt_ = []
            for file_name_ in os.listdir(i):
                txt_.append(i + '\\' + file_name_)
            txt_num = txt_num  # 取多少文件
            txt_ = txt_[-txt_num:]
            num_all = []
            for txt_name in txt_:
                contents = []
                with open(txt_name, 'r') as f:
                    for line in f.readlines():
                        line = line.split('\n')
                        line = line[0].split('\t')
                        line = list(map(float, line))
                        contents.append(line)
                for content in contents:
                    num_all.append("%.4f" % (float(content[1]) / txt_num))
                if len(num_all) > len(contents):
                    for ii in range(len(num_all)):
                        if ii < len(contents):
                            num_all[ii] = "%.4f" % (num_all[ii] + float(num_all[ii + len(contents)]))
                        else:
                            num_all.pop()
                num_all = list(map(float, num_all))
                f.close()
                txt_name_after = newpath + '\\' + file_name_list + '_average\\' + dirs__[a] + "\\" + dirs__[a] + ".txt"
                with open(txt_name_after, "w") as ff:
                    for li in num_all:
                        ff.write(str(li) + "\n")
                ff.close()
            a += 1
        print(file_name_list, "Data reading completed！")
    print("All  data read completed！")


def del_files(path1):
    if os.path.exists(path1):
        shutil.rmtree(path1, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


def remove_bd(newpath):
    path = newpath
    nn_ = []
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        n_p = path + '\\' + file_name_list + '_removebd'
        nn_.append(n_p)
        os.mkdir(n_p)
        root_ = []
        dirs_ = []

        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        root_.pop(0)
        root__ = root_[-1]
        dirs___ = dirs__[-1]
        root_.pop()
        dirs__.pop()
        bd_name = root__ + '\\' + dirs___ + '.txt'
        for i in range(len(root_)):
            data = []
            file_name = root_[i] + '\\' + dirs__[i] + '.txt'
            file_name_ = n_p + '\\' + dirs__[i] + '.txt'
            with open(bd_name) as bd_f:
                bd_file = bd_f.read().split('\n')
            bd_f.close()
            with open(file_name, 'r+') as f:
                file = f.read().split('\n')
            f.close()
            del (bd_file[-1])
            del (file[-1])
            bd_file = list(map(float, bd_file))
            file = list(map(float, file))
            for i in range(len(bd_file)):
                i_num = "%.4f" % ((file[i]) / (bd_file[i]))
                data.append(i_num)
            with open(file_name_, 'w') as f_:
                f_.truncate(0)
                for ii in data:
                    f_.write(ii + "\n")
            f_.close()
    print('Successfully removed the backing!')
    return nn_


def writeinexcel(path, nn):
    lu = []
    path = path
    le_ = 0
    le = 0
    wb1 = xlwt.Workbook(encoding='utf-8')
    w1 = wb1.add_sheet('one')
    ipath_ = path[0]
    file_name_lists_ = []
    file_name_lists = []
    for file_name_ in os.listdir(ipath_):
        file_name_ = re.sub('\D', '', file_name_)
        b = list(file_name_)
        # b.pop()
        ans = "".join(map(str, b))
        file_name_lists_.append(ans)
    file_name_lists_ = list(map(int, file_name_lists_))
    file_name_lists_.sort()
    le = len(file_name_lists_)
    for le_i in range(le):
        w1.write(0, le_i + le_, file_name_lists_[le_i])
    le_ = le_ + len(file_name_lists_)
    for file_name_ in os.listdir(ipath_):
        file_name_lists.append(file_name_)

    # print(file_name_lists)
    # print(nn)
    file_name_lists.sort(key=lambda x: int(x[:-nn]))
    for i_a in range(len(file_name_lists)):
        path_ = ipath_ + '\\' + file_name_lists[i_a]
        ii = 1
        for line in open(path_, encoding='utf-8'):
            if line == '\n':
                continue
            else:
                w1.write(ii, i_a, float(line))
                ii += 1

    wb1.save(path[0] + ".xls")
    p1 = path[0] + ".xls"
    lu.append(p1)

    if len(path) > 1:
        ipath__ = path[1]
        le__ = 0
        lei = 0
        wb2 = xlwt.Workbook(encoding='utf-8')
        w2 = wb2.add_sheet('one')
        file_name_lists_ = []
        file_name_lists = []
        for file_name_ in os.listdir(ipath__):
            file_name_ = re.sub('\D', '', file_name_)
            b = list(file_name_)
            # b.pop()
            ans = "".join(map(str, b))
            file_name_lists_.append(ans)
        file_name_lists_ = list(map(int, file_name_lists_))
        file_name_lists_.sort()
        lei = len(file_name_lists_)
        for le_i in range(lei):
            w2.write(0, le_i + le__, file_name_lists_[le_i])
        le__ = le__ + len(file_name_lists_)
        for file_name_ in os.listdir(ipath__):
            file_name_lists.append(file_name_)
        file_name_lists.sort(key=lambda x: int(x[:-nn]))
        for i_a in range(len(file_name_lists)):
            path_ = ipath__ + '\\' + file_name_lists[i_a]
            ii = 1
            for line in open(path_, encoding='utf-8'):
                if line == '\n':
                    continue
                else:
                    w2.write(ii, i_a, float(line))
                    ii += 1
                    # print(line)
        wb2.save(path[1] + ".xls")
        p2 = path[1] + ".xls"
        lu.append(p2)

    return lu


def koumanbian(x):
    aa = []
    for i in x:
        zz1 = np.polyfit([i for i in range(len(i))], i, 3)  # 拟合
        pp1 = np.poly1d(zz1)
        # aa.append(np.log(i / (pp1([i for i in range(len(i))])))*2) # 先组合在差分的时候乘2
        aa.append(np.log(i / (pp1([i for i in range(len(i))]))))
    return np.array(aa)


def dif_nh3(path):
    # Weak characteristic absorption region 1
    a = 344  # 195.99
    b = 530  # 225.02

    data = pd.read_excel(path)
    columns = data.columns
    wb_all = Workbook()
    wb_all.create_sheet(index=0, title="all")
    ws_all = wb_all.active

    data_line_1 = data.iloc[a - 1:b, :]

    data_line_1 = np.array(data_line_1)

    data_all = data_line_1
    # print(data_all.shape)
    data_all = pd.DataFrame(data_all, columns=columns)
    data_all = data_all.T
    data_all = np.array(data_all)

    data_deal_all_line = koumanbian(data_all)
    deal_data_line_all = pd.DataFrame(data_deal_all_line)
    # print("deal_data_line_all",deal_data_line_all)
    columns_ = deal_data_line_all.columns
    # all_data
    l = len(columns_)
    i1 = 0
    for i in range(l):
        i1 += 1
        lie = deal_data_line_all[columns_[i]]
        for j in range(len(lie)):
            ws_all.cell(1, i + 1, i1)
            ws_all.cell(j + 2, i + 1, lie[j])
    for col_i in range(len(columns)):
        col = columns[col_i]
        ws_all.cell(col_i + 2, l + 1, col)
    ws_all.cell(1, l + 1, l + 1)
    path = path.split('.')[0]
    name = path + '-Sinr-single.xlsx'
    wb_all.save(name)
    path1 = to_pkl(name)
    print("NH3 Data differential completion!")
    return path1


def to_pkl(path):
    # Read excel files
    df1 = DataFrame(pd.read_excel(path))
    dir_name = os.path.dirname(path)
    base_name = os.path.basename(path)
    suffix = base_name.split(".")[0]
    path_ = dir_name + "/" + suffix + ".pkl"
    df1.to_pickle(path_)
    return path_


def get_nh3_data():
    Raw_data_path = "Raw_Sin_NH3_data"
    Processed_data_path = "Test_Processed_Sin_NH3_data"
    txt_num = 300
    numb = 10
    del_files(Processed_data_path)
    read_single_component_data(Raw_data_path, Processed_data_path, txt_num)
    path1 = remove_bd(Processed_data_path)
    path2 = writeinexcel(path1, numb)
    path3 = dif_nh3(path2[0])
    return path3


def index_arrangement(data_x):
    indexed_numbers = list(enumerate(data_x))
    # print("indexed_numbers:",indexed_numbers)
    sorted_numbers = sorted(indexed_numbers, key=lambda x: x[1])
    sorted_values = [x[1] for x in sorted_numbers]  # Sorted values
    sorted_indices = [x[0] for x in sorted_numbers]  # Sorted Indexes
    return sorted_values, sorted_indices


def index_return(sorted_indices, sorted_values_i):
    reverse_index = np.argsort(sorted_indices)[::-1]  # Sorts the list from largest to smallest and returns its index.
    reversed_list = np.take(sorted_values_i, reverse_index)
    reversed_list = reversed_list[::-1]
    return reversed_list


def to_excel(data, data_):
    if os.path.exists("demo.xlsx"):
        os.remove("demo.xlsx")
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    for i in range(len(data)):
        ws.cell(i + 1, 1, data[i])
        ws.cell(i + 1, 2, data_[i])
    wb.save("demo.xlsx")


def to_data(data1, data2, path):
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    for i in range(len(data1[0])):
        ws.cell(1, i + 1, i + 1)
    ws.cell(1, len(data1[0]) + 1, len(data1[0]) + 1)
    for i in range(len(data1)):
        for j in range(len(data1[0])):
            ws.cell(i + 2, j + 1, data1[i][j])
    for jj in range(len(data2)):
        ws.cell(jj + 2, len(data1[0]) + 1, data2[jj])
        if jj == i:
            break

    path = path.split('.')[0]
    name = path + '-Sin-alldata.xlsx'
    wb.save(name)
    path = to_pkl(name)
    return path


def to_sin(data1,data2,path):
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    for i in range(len(data1)):
        ws.cell(1, i + 1, data1[i])
    for j in range(len(data1)):
        ws.cell(2, j + 1, data2[j])
    path = path.split('.')[0]
    name = path + '-Sin-85ppbNH3.xlsx'
    wb.save(name)
    # path = to_pkl(name)
    return path

def to_data_or(data1, data2, path):
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    for i in range(len(data1[0])):
        ws.cell(1, i + 1, i + 1)
    ws.cell(1, len(data1[0]) + 1, len(data1[0]) + 1)
    for i in range(len(data1)):
        for j in range(len(data1[0])):
            ws.cell(i + 2, j + 1, data1[i][j])
    for jj in range(len(data2)):
        ws.cell(jj + 2, len(data1[0]) + 1, data2[jj])
        if jj == i:
            break

    path = path.split('.')[0]
    name = path + '-Sin-alldata.xlsx'
    wb.save(name)
    path = to_pkl(name)
    return path


def find_closest_number(arr, target):
    diff = []
    arr = list(arr)
    for i, num in enumerate(arr):
        diff.append(abs(num - target))
    min_diff = min(diff)
    index = diff.index(min_diff)
    closest_number = arr[index]
    arr.pop(index)
    return closest_number, index


def sin_func(x_data):
    x_max = np.max(abs(x_data))
    x = np.linspace(0, 2 * np.pi, 100 * len(x_data[0]))
    y = x_max * np.sin(x)
    indexs = []
    xvaluelist = []
    yvaluelist = []
    arr = list(y)
    for i in x_data[0]:
        # closest_number, index = find_closest_number(y, i)
        diff = []

        for iii, num in enumerate(arr):
            diff.append(abs(num - i))
        min_diff = min(diff)
        index = diff.index(min_diff)
        closest_number = arr[index]
        arr[index] = 10000000
        # arr.pop(index)
        indexs.append(index)
        yvaluelist.append(closest_number)
    for ii in range(len(indexs)):
        xvaluelist.append(x[indexs[ii]])
    sorted_values, sorted_indices = index_arrangement(indexs)
    indexs_, xvaluelist, yvaluelist = zip(*sorted(zip(indexs, xvaluelist, yvaluelist)))
    return indexs, xvaluelist, yvaluelist, sorted_indices, indexs_


def draw_nh3_plot(x, y, xvaluelist, yvaluelist):
    plt.scatter(x, y, color='black', edgecolors='black')
    plt.plot(xvaluelist, yvaluelist, 'r')
    plt.title("Scatter Plot")
    plt.xlabel("X-axis")
    plt.ylabel("Y-axis")
    plt.ylim(-max(abs(yvaluelist)) - 0.01, max(abs(yvaluelist)) + 0.01)
    plt.show()


def draw_no_plot(x, y, xvaluelist, yvaluelist):
    plt.scatter(x, y, color='black', edgecolors='black')
    plt.plot(xvaluelist, yvaluelist, 'r')
    plt.title("Scatter Plot")
    plt.xlabel("X-axis")
    plt.ylabel("Y-axis")
    plt.ylim(-max(abs(yvaluelist)) - 0.005, max(abs(yvaluelist)) + 0.005)
    plt.show()


def draw_no_plot_all(x, y, xvaluelist, yvaluelist):
    plt.plot(x, y, color='black')
    plt.plot(xvaluelist, yvaluelist, 'r')
    plt.title("Scatter Plot")
    plt.xlabel("X-axis")
    plt.ylabel("Y-axis")
    plt.ylim(-max(abs(yvaluelist)) - 0.005, max(abs(yvaluelist)) + 0.005)
    plt.show()


def draw_nh3_plot_all(x, y, xvaluelist, yvaluelist):
    plt.plot(x, y, color='black')
    plt.plot(xvaluelist, yvaluelist, 'r')
    plt.title("Scatter Plot")
    plt.xlabel("X-axis")
    plt.ylabel("Y-axis")
    plt.ylim(-max(abs(yvaluelist)) - 0.005, max(abs(yvaluelist)) + 0.005)
    plt.show()


def draw_nh3_plot_all_(xvaluelist, yvaluelist):
    plt.plot(xvaluelist, yvaluelist[0], 'r')
    plt.plot(xvaluelist, yvaluelist[1], 'black')
    plt.title("Scatter Plot")
    plt.xlabel("X-axis")
    plt.ylabel("Y-axis")
    # plt.ylim(-max(abs(yvaluelist)) - 0.005, max(abs(yvaluelist)) + 0.005)
    plt.show()


# Define the sine function
def sin_funcc(x, a, b, c):
    return a * np.sin(b * x + c)


def Sin_reconstruction(path):
    data = pd.read_pickle(path)
    data_x = data.values[:, 0:-1]
    data_y = data.values[:, -1]
    path_linear = get_nh3_data()
    data_linear = pd.read_pickle(path_linear)
    data_x_linear = data_linear.values[:, 0:-1]
    data_y_linear = data_linear.values[:, -1]

    # To get the order, x and y have been arranged from smallest to largest.
    indexlist, xvaluelist, yvaluelist, sorted_indices, indexlist_ = sin_func(data_x_linear)
    # to_sin(indexlist, sorted_indices,path_linear)
    # to_sin(xvaluelist, yvaluelist, path_linear)

    y_data = []
    y_data_or = []
    for i in range(len(data_x)):
        # Raw data visualisation
        # draw_nh3_plot(list(range(0, len(data_x[0]), 1)), data_x[i], list(range(0, len(data_x[0]), 1)),
        #                       data_x[i])
        indexs_, yvaluelist_ = zip(*sorted(zip(indexlist, data_x[i])))

        # draw_nh3_plot(xvaluelist, np.array(yvaluelist_), np.linspace(0, 2 * np.pi, len(xvaluelist)),
        #               sin_funcc(np.array(np.linspace(0, 2 * np.pi, len(xvaluelist))), max(abs(np.array(yvaluelist_))), 1, 0))

        popt, pcov = curve_fit(sin_funcc, xvaluelist, yvaluelist_)
        # to_sin(xvaluelist, np.array(yvaluelist_), path_linear)
        sin_y = sin_funcc(np.array(xvaluelist), popt[0], popt[1], popt[2])
        # Visualisation of the fitted data
        to_sin(xvaluelist, sin_y, path_linear)
        print(popt[0], popt[1], popt[2])
        # draw_nh3_plot(xvaluelist, sin_y, np.linspace(0, 2 * np.pi, len(xvaluelist)),
        #           sin_funcc(np.array(np.linspace(0, 2 * np.pi, len(xvaluelist))), popt[0], popt[1], popt[2]))
        # Reducing the fitted data and visualising it
        origin_y = index_return(sorted_indices, sin_y)
        # draw_nh3_plot_all(list(range(0, len(data_x[0]), 1)), data_x[i], list(range(0, len(data_x[0]), 1)),
        #           origin_y)

        # y_data.append(sin_y)
        y_data_or.append(origin_y)




    # draw_nh3_plot_all_(list(range(0, len(data_x[0]), 1)), y_data_or)

    # to_data(y_data,data_y,path_linear)
    to_data_or(y_data_or, data_y, path_linear)
